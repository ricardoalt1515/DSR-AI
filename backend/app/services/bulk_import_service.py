"""Service layer for bulk import staging, review, and finalize."""

from __future__ import annotations

import asyncio
import copy
import csv
import hashlib
import io
import json
import multiprocessing
import re
import tempfile
import time
from collections.abc import Callable, Sequence
from datetime import UTC, datetime, timedelta
from multiprocessing.connection import Connection
from pathlib import Path
from typing import Any
from uuid import UUID

import structlog
from fastapi import HTTPException, status
from PyPDF2 import PdfReader
from sqlalchemy import case, delete, func, select, text, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.bulk_import import ImportItem, ImportRun
from app.models.bulk_import_output import NormalizedLocationDataV1, NormalizedProjectDataV1
from app.models.company import Company
from app.models.location import Location
from app.models.project import Project
from app.models.user import User
from app.schemas.bulk_import import BulkImportFinalizeSummary
from app.services.bulk_import_ai_extractor import (
    BulkImportAIExtractorError,
    ExtractionDiagnostics,
    ParsedRow,
    bulk_import_ai_extractor,
)
from app.services.s3_service import download_file_content
from app.services.storage_delete_service import delete_storage_keys
from app.templates.assessment_questionnaire import get_assessment_questionnaire

logger = structlog.get_logger(__name__)

MAX_PROCESSING_ATTEMPTS = 3
LEASE_SECONDS = 300
RETRY_BASE_SECONDS = 30
RETRY_MAX_SECONDS = 600
RETRY_JITTER_PCT = 0.2
PROCESSING_ERROR_MAX_LENGTH = 500

MAX_IMPORT_FILE_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 2000
MAX_IMPORT_CELLS = 30000
MAX_IMPORT_ITEMS = 4000
PARSER_TIMEOUT_SECONDS = 25
MAX_TEXT_LEN = 4000
PARSER_WORKER_NAME = "bulk-import-parse-worker"

RETENTION_DAYS_UNFINALIZED = 90

ALLOWED_BULK_IMPORT_EXTENSIONS = {
    (ext if ext.startswith(".") else f".{ext}").casefold()
    for ext in settings.bulk_import_allowed_extensions_list
}


class ParserLimitError(ValueError):
    """Raised when parser limits are exceeded."""


def _normalize_token(value: str | None) -> str:
    if not value:
        return ""
    normalized = value.strip().casefold()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return " ".join(normalized.split())


def _sanitize_text(value: str | None, max_len: int = MAX_TEXT_LEN) -> str | None:
    if value is None:
        return None
    cleaned = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]", "", value)
    cleaned = cleaned.strip()
    if len(cleaned) > max_len:
        return cleaned[:max_len]
    return cleaned


def _sanitize_payload(value: Any) -> Any:
    if isinstance(value, str):
        return _sanitize_text(value)
    if isinstance(value, dict):
        sanitized: dict[str, Any] = {}
        for key, nested_value in value.items():
            key_text = _sanitize_text(str(key), max_len=100)
            if not key_text:
                continue
            sanitized[key_text] = _sanitize_payload(nested_value)
        return sanitized
    if isinstance(value, list):
        return [_sanitize_payload(item) for item in value]
    return value


def _dedupe_backoff_seconds(run_id: UUID, attempt: int) -> int:
    base = min(RETRY_BASE_SECONDS * (2 ** max(attempt - 1, 0)), RETRY_MAX_SECONDS)
    digest = hashlib.sha256(f"{run_id}:{attempt}".encode()).digest()
    jitter_value = int.from_bytes(digest[:8], "big") / 2**64
    jitter_factor = (jitter_value * 2.0 - 1.0) * RETRY_JITTER_PCT
    backoff = round(base * (1 + jitter_factor))
    return max(1, min(backoff, RETRY_MAX_SECONDS))


def _truncate_error(error: str) -> str:
    return error[:PROCESSING_ERROR_MAX_LENGTH]


def _parse_source_subprocess_entrypoint(
    parse_callable: Callable[[str, bytes], list[ParsedRow]],
    filename: str,
    file_bytes: bytes,
    child_conn: Connection,
) -> None:
    """Parse in subprocess, return payload via pipe.

    Payload format:
    - ("ok_file", temp_path)
    - ("error", exc_type_name, message)
    """
    try:
        parsed = parse_callable(filename, file_bytes)
        serialized = _serialize_parsed_rows(parsed)
        with tempfile.NamedTemporaryFile(
            mode="w",
            prefix="bulk-import-parse-",
            suffix=".json",
            encoding="utf-8",
            delete=False,
        ) as temp_file:
            json.dump(serialized, temp_file, separators=(",", ":"))
            temp_path = temp_file.name
        child_conn.send(("ok_file", temp_path))
    except Exception as exc:  # pragma: no cover - validated by parent behavior tests
        child_conn.send(("error", type(exc).__name__, str(exc)))
    finally:
        child_conn.close()


def _default_parse_callable(filename: str, file_bytes: bytes) -> list[ParsedRow]:
    parser_service = BulkImportService()
    return parser_service._parse_source(filename, file_bytes)


def _serialize_parsed_rows(rows: list[ParsedRow]) -> list[dict[str, Any]]:
    serialized: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, ParsedRow):
            raise ValueError("parser_invalid_result")
        serialized.append(
            {
                "location_data": row.location_data,
                "project_data": row.project_data,
                "raw": row.raw,
            }
        )
    return serialized


def _coerce_str_dict(value: Any, *, allow_none: bool) -> dict[str, str] | None:
    if value is None:
        if allow_none:
            return None
        raise ValueError("parser_invalid_result")
    if not isinstance(value, dict):
        raise ValueError("parser_invalid_result")
    coerced: dict[str, str] = {}
    for key, nested in value.items():
        if not isinstance(key, str):
            raise ValueError("parser_invalid_result")
        if nested is None:
            coerced[key] = ""
            continue
        if not isinstance(nested, str):
            raise ValueError("parser_invalid_result")
        coerced[key] = nested
    return coerced


def _deserialize_parsed_rows(payload: Any) -> list[ParsedRow]:
    if not isinstance(payload, list):
        raise ValueError("parser_invalid_result")

    rows: list[ParsedRow] = []
    for item in payload:
        if not isinstance(item, dict):
            raise ValueError("parser_invalid_result")
        location_data = _coerce_str_dict(item.get("location_data"), allow_none=True)
        project_data = _coerce_str_dict(item.get("project_data"), allow_none=True)
        raw_data = _coerce_str_dict(item.get("raw"), allow_none=False)
        if raw_data is None:
            raise ValueError("parser_invalid_result")
        rows.append(
            ParsedRow(
                location_data=location_data,
                project_data=project_data,
                raw=raw_data,
            )
        )
    return rows


def _load_json_parse_result(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as temp_file:
        return json.load(temp_file)


def _log_diagnostics(diagnostics: ExtractionDiagnostics | None) -> dict[str, object]:
    if diagnostics is None:
        return {}
    return {
        "route": diagnostics.route,
        "char_count": diagnostics.char_count,
        "truncated": diagnostics.truncated,
    }


class BulkImportService:
    """Bulk import orchestration across worker and API layers."""

    async def claim_next_run(self, db: AsyncSession) -> ImportRun | None:
        candidate = (
            select(ImportRun.id)
            .where(ImportRun.status == "uploaded")
            .where(ImportRun.processing_attempts < MAX_PROCESSING_ATTEMPTS)
            .where(
                (ImportRun.processing_available_at.is_(None))
                | (ImportRun.processing_available_at <= func.now())
            )
            .order_by(ImportRun.processing_available_at.nullsfirst(), ImportRun.created_at)
            .with_for_update(skip_locked=True)
            .limit(1)
            .cte("candidate")
        )
        stmt = (
            update(ImportRun)
            .where(ImportRun.id.in_(select(candidate.c.id)))
            .where(ImportRun.status == "uploaded")
            .values(
                status="processing",
                progress_step="reading_file",
                processing_attempts=ImportRun.processing_attempts + 1,
                processing_started_at=func.now(),
                processing_available_at=func.now() + text(f"INTERVAL '{LEASE_SECONDS} seconds'"),
                processing_error=None,
            )
            .returning(ImportRun)
        )
        result = await db.execute(stmt)
        run = result.scalar_one_or_none()
        return run

    async def requeue_stale_runs(self, db: AsyncSession, limit: int = 100) -> int:
        stale_requeue = (
            select(ImportRun.id)
            .where(ImportRun.status == "processing")
            .where(ImportRun.processing_available_at < func.now())
            .where(ImportRun.processing_attempts < MAX_PROCESSING_ATTEMPTS)
            .order_by(ImportRun.processing_available_at)
            .limit(limit)
            .cte("stale_requeue")
        )
        requeue_stmt = (
            update(ImportRun)
            .where(ImportRun.id.in_(select(stale_requeue.c.id)))
            .where(ImportRun.status == "processing")
            .values(
                status="uploaded",
                processing_error="lease_expired_requeued",
                processing_available_at=func.now(),
                processing_started_at=None,
            )
            .returning(ImportRun.id)
        )
        requeued_result = await db.execute(requeue_stmt)
        return len(requeued_result.scalars().all())

    async def fail_exhausted_runs(self, db: AsyncSession, limit: int = 100) -> int:
        exhausted = (
            select(ImportRun.id)
            .where(ImportRun.status.in_(["uploaded", "processing"]))
            .where(ImportRun.processing_attempts >= MAX_PROCESSING_ATTEMPTS)
            .order_by(ImportRun.updated_at)
            .limit(limit)
            .cte("exhausted")
        )
        stmt = (
            update(ImportRun)
            .where(ImportRun.id.in_(select(exhausted.c.id)))
            .where(ImportRun.status.in_(["uploaded", "processing"]))
            .values(
                status="failed",
                progress_step=None,
                processing_error="max_attempts_reached",
            )
            .returning(ImportRun.id)
        )
        result = await db.execute(stmt)
        return len(result.scalars().all())

    async def purge_expired_artifacts(self, db: AsyncSession, limit: int = 100) -> int:
        cutoff = datetime.now(UTC) - timedelta(days=RETENTION_DAYS_UNFINALIZED)
        result = await db.execute(
            select(ImportRun)
            .where(ImportRun.created_at < cutoff)
            .where(
                ImportRun.status.in_(
                    ["uploaded", "processing", "review_ready", "failed", "no_data"]
                )
            )
            .where(ImportRun.artifacts_purged_at.is_(None))
            .order_by(ImportRun.created_at)
            .limit(limit)
            .with_for_update(skip_locked=True)
        )
        runs = result.scalars().all()
        if not runs:
            return 0

        purged = 0
        for run in runs:
            try:
                await delete_storage_keys([run.source_file_path])
            except Exception:
                logger.warning(
                    "bulk_import_purge_storage_failed", run_id=str(run.id), exc_info=True
                )
                continue
            run.artifacts_purged_at = datetime.now(UTC)
            run.source_file_path = "imports/purged"
            run.processing_error = None
            await db.execute(
                update(ImportItem)
                .where(ImportItem.run_id == run.id)
                .values(
                    extracted_data={},
                    user_amendments=None,
                    review_notes=None,
                )
            )
            purged += 1
        return purged

    async def process_run(self, db: AsyncSession, run: ImportRun) -> None:
        ai_call_duration_ms: float | None = None
        run_id = run.id
        try:
            if run.status != "processing":
                raise ValueError("run_not_processing")

            await self._persist_progress_checkpoint(db, run, "reading_file")
            file_bytes = await download_file_content(run.source_file_path)
            if not file_bytes:
                raise ValueError("empty_file")
            if len(file_bytes) > MAX_IMPORT_FILE_BYTES:
                raise ParserLimitError("max_file_size_exceeded")

            await self._persist_progress_checkpoint(db, run, "identifying_locations")
            extension = Path(run.source_filename).suffix.casefold()
            if extension not in ALLOWED_BULK_IMPORT_EXTENSIONS:
                raise ValueError("unsupported_file_type")

            await self._persist_progress_checkpoint(db, run, "extracting_streams")
            ai_started = time.perf_counter()
            try:
                extraction_result = await bulk_import_ai_extractor.extract_parsed_rows(
                    file_bytes=file_bytes,
                    filename=run.source_filename,
                )
                ai_call_duration_ms = round((time.perf_counter() - ai_started) * 1000, 2)
                logger.info(
                    "bulk_import_api_status",
                    run_id=str(run.id),
                    filename=run.source_filename,
                    status="success",
                    duration_ms=ai_call_duration_ms,
                    **_log_diagnostics(extraction_result.diagnostics),
                )
                parsed_rows = extraction_result.rows
            except BulkImportAIExtractorError as exc:
                ai_call_duration_ms = round((time.perf_counter() - ai_started) * 1000, 2)
                logger.info(
                    "bulk_import_api_status",
                    run_id=str(run.id),
                    filename=run.source_filename,
                    status="failed",
                    duration_ms=ai_call_duration_ms,
                    error_code=exc.code,
                    **_log_diagnostics(exc.diagnostics),
                )
                raise ValueError(exc.code) from exc

            await self._persist_progress_checkpoint(db, run, "categorizing")
            await db.execute(delete(ImportItem).where(ImportItem.run_id == run.id))
            staged_items = await self._build_import_items(db, run, parsed_rows)

            if not staged_items:
                run.status = "no_data"
                run.progress_step = None
                run.total_items = 0
                run.accepted_count = 0
                run.rejected_count = 0
                run.amended_count = 0
                run.invalid_count = 0
                run.duplicate_count = 0
                run.processing_error = None
                await db.flush()
                return

            if len(staged_items) > MAX_IMPORT_ITEMS:
                raise ParserLimitError("max_items_exceeded")

            db.add_all(staged_items)
            await db.flush()
            await self.refresh_run_counters(db, run)
            run.status = "review_ready"
            run.progress_step = None
            run.processing_error = None
            await db.flush()
        except Exception as exc:
            await db.rollback()
            await self._handle_processing_failure(db, run_id=run_id, exc=exc)

    async def _persist_progress_checkpoint(
        self,
        db: AsyncSession,
        run: ImportRun,
        phase: str,
    ) -> None:
        run.progress_step = phase
        await db.flush()
        await db.commit()

    async def finalize_run(
        self,
        db: AsyncSession,
        *,
        run_id: UUID,
        organization_id: UUID,
        current_user: User,
    ) -> BulkImportFinalizeSummary:
        result = await db.execute(
            select(ImportRun)
            .where(ImportRun.id == run_id, ImportRun.organization_id == organization_id)
            .with_for_update()
        )
        run = result.scalar_one_or_none()
        if not run:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Run not found")

        if run.status == "completed":
            return self._summary_from_run(run)
        if run.status == "finalizing":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT, detail="Run already finalizing"
            )
        if run.status != "review_ready":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run is not ready for finalize",
            )

        await self._assert_finalize_ready(db, run)

        run.status = "finalizing"
        run.progress_step = "finalizing"
        run.finalized_by_user_id = current_user.id
        await db.flush()

        all_items_result = await db.execute(
            select(ImportItem).where(ImportItem.run_id == run.id).order_by(ImportItem.created_at)
        )
        all_items = all_items_result.scalars().all()

        active_items = [item for item in all_items if item.status in {"accepted", "amended"}]
        location_items = [item for item in active_items if item.item_type == "location"]
        project_items = [item for item in active_items if item.item_type == "project"]

        await self._assert_finalize_no_new_live_duplicates(
            db=db,
            run=run,
            active_items=active_items,
        )

        company = await self._load_entrypoint_company(db, run)
        company_cache: dict[UUID, Company] = {}
        if company is not None:
            company_cache[company.id] = company

        location_by_parent_item_id: dict[UUID, Location] = {}
        created_locations_count = 0
        created_projects_count = 0

        for item in location_items:
            normalized = self._validated_location_data(item)
            if company is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Location items are not allowed for location entrypoint",
                )
            location = Location(
                organization_id=run.organization_id,
                company_id=company.id,
                name=normalized.name,
                city=normalized.city,
                state=normalized.state,
                address=normalized.address,
                created_by_user_id=current_user.id,
            )
            db.add(location)
            await db.flush()
            item.created_location_id = location.id
            location_by_parent_item_id[item.id] = location
            created_locations_count += 1

        entrypoint_location: Location | None = None
        if run.entrypoint_type == "location":
            entrypoint_location = await db.get(Location, run.entrypoint_id)
            if (
                not entrypoint_location
                or entrypoint_location.organization_id != run.organization_id
            ):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND, detail="Entrypoint location not found"
                )

        for item in project_items:
            normalized = self._validated_project_data(item)
            target_location = await self._resolve_project_location_for_finalize(
                db=db,
                run=run,
                item=item,
                location_by_parent_item_id=location_by_parent_item_id,
                fallback_location=entrypoint_location,
            )
            company_for_project = company_cache.get(target_location.company_id)
            if company_for_project is None:
                company_for_project = await db.get(Company, target_location.company_id)
                if company_for_project:
                    company_cache[target_location.company_id] = company_for_project
            if not company_for_project:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Target location has no company",
                )

            project_data: dict[str, object] = {
                "technical_sections": copy.deepcopy(get_assessment_questionnaire())
            }
            if normalized.category and normalized.category.strip():
                project_data["bulk_import_category"] = normalized.category.strip()

            project = Project(
                organization_id=run.organization_id,
                user_id=current_user.id,
                location_id=target_location.id,
                name=normalized.name,
                client=company_for_project.name,
                sector=normalized.sector or company_for_project.sector,
                subsector=normalized.subsector or company_for_project.subsector,
                location=f"{target_location.name}, {target_location.city}",
                project_type=normalized.project_type,
                description=normalized.description,
                budget=0.0,
                schedule_summary="To be defined",
                tags=[],
                status="In Preparation",
                progress=0,
                project_data=project_data,
            )
            db.add(project)
            await db.flush()
            item.created_project_id = project.id
            created_projects_count += 1

        rejected_count = sum(1 for item in all_items if item.status == "rejected")
        invalid_count = sum(1 for item in all_items if item.status == "invalid")
        duplicates_resolved = sum(
            1
            for item in active_items
            if item.duplicate_candidates
            and len(item.duplicate_candidates) > 0
            and item.confirm_create_new
        )

        summary = BulkImportFinalizeSummary(
            run_id=run.id,
            locations_created=created_locations_count,
            projects_created=created_projects_count,
            rejected=rejected_count,
            invalid=invalid_count,
            duplicates_resolved=duplicates_resolved,
        )

        run.status = "completed"
        run.progress_step = None
        run.finalized_at = datetime.now(UTC)
        run.summary_data = summary.model_dump(mode="json")
        await self.refresh_run_counters(db, run)

        try:
            await delete_storage_keys([run.source_file_path])
            run.artifacts_purged_at = datetime.now(UTC)
        except Exception:
            logger.warning(
                "bulk_import_finalize_artifact_delete_failed", run_id=str(run.id), exc_info=True
            )

        await db.flush()
        return summary

    async def _assert_finalize_no_new_live_duplicates(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        active_items: Sequence[ImportItem],
    ) -> None:
        if not active_items:
            return

        company: Company | None = None
        if run.entrypoint_type == "company":
            company = await self._load_entrypoint_company(db, run)

        entrypoint_location: Location | None = None
        if run.entrypoint_type == "location":
            entrypoint_location = await db.get(Location, run.entrypoint_id)
            if (
                not entrypoint_location
                or entrypoint_location.organization_id != run.organization_id
            ):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Entrypoint location not found",
                )

        active_items_by_id = {item.id: item for item in active_items}

        for item in active_items:
            if item.confirm_create_new:
                continue

            if item.item_type == "location":
                if company is None:
                    continue
                normalized_location = self._validated_location_data(item)
                location_duplicate_exists = await self._location_duplicate_exists_for_finalize(
                    db=db,
                    organization_id=run.organization_id,
                    company_id=company.id,
                    location_data=normalized_location,
                )
                if location_duplicate_exists:
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail=f"Duplicate detected before finalize for item {item.id}",
                    )
                continue

            normalized_project = self._validated_project_data(item)
            target_location_id = await self._project_finalize_location_id_for_duplicate_recheck(
                db=db,
                run=run,
                item=item,
                active_items_by_id=active_items_by_id,
                entrypoint_location=entrypoint_location,
                company=company,
            )
            if target_location_id is None:
                continue

            project_duplicate_exists = await self._project_duplicate_exists_for_finalize(
                db=db,
                organization_id=run.organization_id,
                location_id=target_location_id,
                project_name=normalized_project.name,
            )
            if project_duplicate_exists:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Duplicate detected before finalize for item {item.id}",
                )

    async def _location_duplicate_exists_for_finalize(
        self,
        *,
        db: AsyncSession,
        organization_id: UUID,
        company_id: UUID,
        location_data: NormalizedLocationDataV1,
    ) -> bool:
        result = await db.execute(
            select(Location.id)
            .where(
                Location.organization_id == organization_id,
                Location.company_id == company_id,
                func.lower(Location.name) == location_data.name.casefold(),
                func.lower(Location.city) == location_data.city.casefold(),
                func.lower(Location.state) == location_data.state.casefold(),
            )
            .limit(1)
        )
        return result.scalar_one_or_none() is not None

    async def _project_duplicate_exists_for_finalize(
        self,
        *,
        db: AsyncSession,
        organization_id: UUID,
        location_id: UUID,
        project_name: str,
    ) -> bool:
        normalized_project_name = _sanitize_text(project_name)
        if not normalized_project_name:
            return False
        result = await db.execute(
            select(Project.id)
            .where(
                Project.organization_id == organization_id,
                Project.location_id == location_id,
                func.lower(Project.name) == normalized_project_name.casefold(),
            )
            .limit(1)
        )
        return result.scalar_one_or_none() is not None

    async def _project_finalize_location_id_for_duplicate_recheck(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        active_items_by_id: dict[UUID, ImportItem],
        entrypoint_location: Location | None,
        company: Company | None,
    ) -> UUID | None:
        if run.entrypoint_type == "location":
            if entrypoint_location is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Entrypoint location missing",
                )
            return entrypoint_location.id

        if item.parent_item_id is not None:
            parent_item = active_items_by_id.get(item.parent_item_id)
            if parent_item is not None:
                return None

        location_name = _sanitize_text(str(item.normalized_data.get("location_name") or ""))
        city = _sanitize_text(str(item.normalized_data.get("location_city") or ""))
        state_value = _sanitize_text(str(item.normalized_data.get("location_state") or ""))
        if not location_name or not city or not state_value:
            return None

        if company is None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Company entrypoint required",
            )

        result = await db.execute(
            select(Location.id)
            .where(
                Location.organization_id == run.organization_id,
                Location.company_id == company.id,
                func.lower(Location.name) == location_name.casefold(),
                func.lower(Location.city) == city.casefold(),
                func.lower(Location.state) == state_value.casefold(),
            )
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def refresh_run_counters(self, db: AsyncSession, run: ImportRun) -> None:
        counts_result = await db.execute(
            select(
                func.count(ImportItem.id),
                func.sum(case((ImportItem.status == "accepted", 1), else_=0)),
                func.sum(case((ImportItem.status == "rejected", 1), else_=0)),
                func.sum(case((ImportItem.status == "amended", 1), else_=0)),
                func.sum(case((ImportItem.status == "invalid", 1), else_=0)),
                func.sum(
                    case(
                        (
                            func.jsonb_typeof(ImportItem.duplicate_candidates) == "array",
                            case(
                                (func.jsonb_array_length(ImportItem.duplicate_candidates) > 0, 1),
                                else_=0,
                            ),
                        ),
                        else_=0,
                    )
                ),
            ).where(ImportItem.run_id == run.id)
        )
        counts = counts_result.one()
        run.total_items = int(counts[0] or 0)
        run.accepted_count = int(counts[1] or 0)
        run.rejected_count = int(counts[2] or 0)
        run.amended_count = int(counts[3] or 0)
        run.invalid_count = int(counts[4] or 0)
        run.duplicate_count = int(counts[5] or 0)

    async def update_item_decision(
        self,
        db: AsyncSession,
        *,
        organization_id: UUID,
        item_id: UUID,
        action: str,
        normalized_data: dict[str, object] | None,
        review_notes: str | None,
        confirm_create_new: bool | None,
    ) -> ImportItem:
        run_result = await db.execute(
            select(ImportRun)
            .join(ImportItem, ImportItem.run_id == ImportRun.id)
            .where(ImportItem.id == item_id, ImportItem.organization_id == organization_id)
            .where(ImportRun.organization_id == organization_id)
            .with_for_update()
        )
        run = run_result.scalar_one_or_none()
        if not run:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")

        result = await db.execute(
            select(ImportItem)
            .where(ImportItem.id == item_id, ImportItem.organization_id == organization_id)
            .where(ImportItem.run_id == run.id)
            .with_for_update()
        )
        item = result.scalar_one_or_none()
        if not item:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
        if run.status != "review_ready":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run is not editable",
            )

        if item.status == "invalid":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Invalid items are terminal",
            )

        if review_notes is not None:
            item.review_notes = _sanitize_text(review_notes, max_len=1000)
        if confirm_create_new is not None:
            item.confirm_create_new = confirm_create_new

        if action == "accept":
            self._ensure_duplicate_confirmation(item, target_status="accepted")
            item.status = "accepted"
            item.needs_review = self._needs_review(item.item_type, item.normalized_data)
        elif action == "amend":
            if normalized_data is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="normalized_data required for amend",
                )
            sanitized_patch = _sanitize_payload(normalized_data)
            merged = dict(item.normalized_data)
            merged.update(sanitized_patch)
            item.normalized_data = merged
            item.user_amendments = sanitized_patch
            self._ensure_duplicate_confirmation(item, target_status="amended")
            item.status = "amended"
            item.needs_review = self._needs_review(item.item_type, item.normalized_data)
        elif action == "reject":
            item.status = "rejected"
            item.needs_review = False
            if item.item_type == "location":
                await db.execute(
                    update(ImportItem)
                    .where(ImportItem.parent_item_id == item.id)
                    .where(ImportItem.run_id == item.run_id)
                    .values(status="rejected", needs_review=False)
                )
        elif action == "reset":
            item.status = "pending_review"
            item.needs_review = True
            item.confirm_create_new = False
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid action")

        await self.refresh_run_counters(db, run)
        await db.flush()
        return item

    async def get_run(
        self, db: AsyncSession, *, organization_id: UUID, run_id: UUID
    ) -> ImportRun | None:
        result = await db.execute(
            select(ImportRun)
            .where(ImportRun.id == run_id, ImportRun.organization_id == organization_id)
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def get_item(
        self, db: AsyncSession, *, organization_id: UUID, item_id: UUID
    ) -> ImportItem | None:
        result = await db.execute(
            select(ImportItem)
            .where(ImportItem.id == item_id, ImportItem.organization_id == organization_id)
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def list_items(
        self,
        db: AsyncSession,
        *,
        organization_id: UUID,
        run_id: UUID,
        page: int,
        size: int,
        status_filter: str | None,
    ) -> tuple[list[ImportItem], int]:
        query = select(ImportItem).where(
            ImportItem.run_id == run_id,
            ImportItem.organization_id == organization_id,
        )
        if status_filter:
            query = query.where(ImportItem.status == status_filter)

        count_result = await db.execute(select(func.count()).select_from(query.subquery()))
        total = int(count_result.scalar_one() or 0)

        paged_result = await db.execute(
            query.order_by(ImportItem.created_at, ImportItem.id)
            .offset((page - 1) * size)
            .limit(size)
        )
        return list(paged_result.scalars().all()), total

    def _summary_from_run(self, run: ImportRun) -> BulkImportFinalizeSummary:
        if not run.summary_data:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run completed without summary",
            )
        return BulkImportFinalizeSummary.model_validate(run.summary_data)

    def get_run_summary(self, run: ImportRun) -> BulkImportFinalizeSummary:
        return self._summary_from_run(run)

    async def _assert_finalize_ready(self, db: AsyncSession, run: ImportRun) -> None:
        pending_count_result = await db.execute(
            select(func.count(ImportItem.id)).where(
                ImportItem.run_id == run.id,
                ImportItem.status == "pending_review",
            )
        )
        pending_count = int(pending_count_result.scalar_one() or 0)
        if pending_count > 0:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Run has pending_review items",
            )

        items_result = await db.execute(select(ImportItem).where(ImportItem.run_id == run.id))
        items = items_result.scalars().all()
        by_id = {item.id: item for item in items}

        for item in items:
            if item.status not in {"accepted", "amended"}:
                continue
            self._ensure_duplicate_confirmation(item)
            if item.needs_review:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Item {item.id} still needs review",
                )
            if item.item_type == "location":
                self._validated_location_data(item)
                if run.entrypoint_type == "location":
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail="Location items not allowed for location entrypoint",
                    )
            else:
                self._validated_project_data(item)
                if item.parent_item_id:
                    parent = by_id.get(item.parent_item_id)
                    if parent is None or parent.item_type != "location":
                        raise HTTPException(
                            status_code=status.HTTP_409_CONFLICT,
                            detail="Project item parent invalid",
                        )
                    if parent.status not in {"accepted", "amended"}:
                        raise HTTPException(
                            status_code=status.HTTP_409_CONFLICT,
                            detail="Project item parent is not accepted",
                        )

    async def _resolve_project_location_for_finalize(
        self,
        *,
        db: AsyncSession,
        run: ImportRun,
        item: ImportItem,
        location_by_parent_item_id: dict[UUID, Location],
        fallback_location: Location | None,
    ) -> Location:
        if run.entrypoint_type == "location":
            if fallback_location is None:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Entrypoint location missing",
                )
            return fallback_location

        if item.parent_item_id and item.parent_item_id in location_by_parent_item_id:
            return location_by_parent_item_id[item.parent_item_id]

        location_name = _sanitize_text(str(item.normalized_data.get("location_name") or ""))
        city = _sanitize_text(str(item.normalized_data.get("location_city") or ""))
        state_value = _sanitize_text(str(item.normalized_data.get("location_state") or ""))
        if not location_name or not city or not state_value:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Project item location unresolved",
            )

        company = await self._load_entrypoint_company(db, run)
        if company is None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Company entrypoint required",
            )

        result = await db.execute(
            select(Location).where(
                Location.organization_id == run.organization_id,
                Location.company_id == company.id,
                func.lower(Location.name) == location_name.casefold(),
                func.lower(Location.city) == city.casefold(),
                func.lower(Location.state) == state_value.casefold(),
            )
        )
        location = result.scalar_one_or_none()
        if not location:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Project item location not found",
            )
        return location

    def _ensure_duplicate_confirmation(
        self, item: ImportItem, target_status: str | None = None
    ) -> None:
        status_to_validate = target_status or item.status
        if status_to_validate not in {"accepted", "amended"}:
            return
        candidates = item.duplicate_candidates or []
        if len(candidates) > 0 and not item.confirm_create_new:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Duplicate requires confirm_create_new=true",
            )

    def _validated_location_data(self, item: ImportItem) -> NormalizedLocationDataV1:
        try:
            return NormalizedLocationDataV1.model_validate(item.normalized_data)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Location item invalid: {item.id}",
            ) from exc

    def _validated_project_data(self, item: ImportItem) -> NormalizedProjectDataV1:
        try:
            data = NormalizedProjectDataV1.model_validate(item.normalized_data)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Project item invalid: {item.id}",
            ) from exc
        return data

    async def _load_entrypoint_company(self, db: AsyncSession, run: ImportRun) -> Company | None:
        if run.entrypoint_type != "company":
            return None
        company = await db.get(Company, run.entrypoint_id)
        if not company or company.organization_id != run.organization_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Entrypoint company not found"
            )
        return company

    def _needs_review(self, item_type: str, payload: dict[str, object]) -> bool:
        if item_type == "location":
            fields = [
                _sanitize_text(str(payload.get("name") or "")),
                _sanitize_text(str(payload.get("city") or "")),
                _sanitize_text(str(payload.get("state") or "")),
            ]
            return any(not field for field in fields)
        fields = [_sanitize_text(str(payload.get("name") or ""))]
        return any(not field for field in fields)

    def _needs_review_by_confidence(self, confidence: int) -> bool:
        return confidence < 80

    def _confidence_from_raw(self, raw: dict[str, str], key: str, default: int) -> int:
        value = raw.get(key)
        if not value:
            return default
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            return default
        if parsed < 0:
            return 0
        if parsed > 100:
            return 100
        return parsed

    async def _handle_processing_failure(
        self,
        db: AsyncSession,
        run_id: UUID,
        exc: Exception,
    ) -> None:
        retryable = not isinstance(exc, (ParserLimitError, ValueError))
        reason = _truncate_error(str(exc) or "processing_failed")

        run_result = await db.execute(
            select(ImportRun).where(ImportRun.id == run_id).with_for_update()
        )
        run = run_result.scalar_one_or_none()
        if run is None:
            return

        if retryable and run.processing_attempts < MAX_PROCESSING_ATTEMPTS:
            run.status = "uploaded"
            run.progress_step = None
            run.processing_error = reason
            run.processing_available_at = datetime.now(UTC) + timedelta(
                seconds=_dedupe_backoff_seconds(run.id, run.processing_attempts)
            )
            run.processing_started_at = None
            await db.flush()
            return

        run.status = "failed"
        run.progress_step = None
        run.processing_error = reason
        await db.flush()

    async def _parse_source_with_hard_timeout(
        self,
        *,
        filename: str,
        file_bytes: bytes,
        timeout_seconds: float,
        parse_callable: Callable[[str, bytes], list[ParsedRow]] | None = None,
    ) -> list[ParsedRow]:
        """Run parser in killable subprocess; enforce real timeout cancelation."""
        parser = parse_callable or _default_parse_callable
        ctx = multiprocessing.get_context("spawn")
        parent_conn, child_conn = ctx.Pipe(duplex=False)
        process = ctx.Process(
            name=PARSER_WORKER_NAME,
            target=_parse_source_subprocess_entrypoint,
            args=(parser, filename, file_bytes, child_conn),
            daemon=True,
        )

        process.start()
        child_conn.close()
        payload: tuple[object, ...] | None = None
        timed_out = False
        try:
            deadline = time.monotonic() + timeout_seconds
            while True:
                remaining = deadline - time.monotonic()
                if remaining <= 0:
                    timed_out = True
                    break

                if parent_conn.poll(min(0.1, remaining)):
                    try:
                        payload = parent_conn.recv()
                    except EOFError:
                        payload = None
                    break

                if process.exitcode is not None and not parent_conn.poll():
                    break

            if payload is None:
                if process.is_alive():
                    process.terminate()
                    await asyncio.to_thread(process.join, 1.0)
                    if process.is_alive():
                        process.kill()
                        await asyncio.to_thread(process.join, 1.0)
                if timed_out:
                    raise ValueError("parser_timeout")
                raise ValueError("parser_no_result")

            await asyncio.to_thread(process.join, 1.0)
            if process.is_alive():
                process.terminate()
                await asyncio.to_thread(process.join, 1.0)
                if process.is_alive():
                    process.kill()
                    await asyncio.to_thread(process.join, 1.0)
        finally:
            parent_conn.close()

        status_code = payload[0]
        if status_code == "ok_file":
            temp_path = payload[1]
            if not isinstance(temp_path, str):
                raise ValueError("parser_invalid_result_path")
            temp_file_path = Path(temp_path)
            try:
                serialized = await asyncio.to_thread(_load_json_parse_result, temp_file_path)
            finally:
                try:
                    await asyncio.to_thread(temp_file_path.unlink, missing_ok=True)
                except OSError:
                    logger.warning("bulk_import_parse_tempfile_cleanup_failed", path=temp_path)
            return _deserialize_parsed_rows(serialized)

        if status_code == "error":
            error_type = payload[1]
            error_message = payload[2]
            if error_type == "ParserLimitError":
                raise ParserLimitError(error_message)
            if error_type == "ValueError":
                raise ValueError(error_message)
            raise RuntimeError(error_message)

        raise ValueError("parser_invalid_payload")

    def _parse_source(self, filename: str, file_bytes: bytes) -> list[ParsedRow]:
        extension = Path(filename).suffix.casefold()
        if extension == ".csv":
            rows = self._parse_csv(file_bytes)
        elif extension == ".xlsx":
            rows = self._parse_excel(file_bytes)
        elif extension == ".xls":
            raise ValueError("legacy_xls_not_supported")
        elif extension == ".pdf":
            rows = self._parse_pdf(file_bytes)
        else:
            raise ValueError("unsupported_file_type")

        parsed: list[ParsedRow] = []
        for row in rows:
            normalized_row = {
                str(k).strip().casefold(): _sanitize_text(str(v) if v is not None else "")
                for k, v in row.items()
            }
            location_data = self._extract_location_data(normalized_row)
            project_data = self._extract_project_data(normalized_row)
            if not location_data and not project_data:
                continue
            parsed.append(
                ParsedRow(
                    location_data=location_data,
                    project_data=project_data,
                    raw={k: v or "" for k, v in normalized_row.items()},
                )
            )
        return parsed

    def _parse_csv(self, file_bytes: bytes) -> list[dict[str, str]]:
        text_value = file_bytes.decode("utf-8-sig", errors="ignore")
        return self._parse_delimited_text(text_value)

    def _parse_excel(self, file_bytes: bytes) -> list[dict[str, str]]:
        load_workbook = self._get_openpyxl_load_workbook()
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            sheets = workbook.worksheets
            if not sheets:
                return []
            sheet = sheets[0]
            rows_iter = sheet.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return []
            headers = [str(value).strip() if value is not None else "" for value in header_row]

            rows: list[dict[str, str]] = []
            row_count = 0
            cell_count = 0
            for values in rows_iter:
                row_count += 1
                if row_count > MAX_IMPORT_ROWS:
                    raise ParserLimitError("max_rows_exceeded")
                cell_count += len(values)
                if cell_count > MAX_IMPORT_CELLS:
                    raise ParserLimitError("max_cells_exceeded")
                row: dict[str, str] = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    value = values[index] if index < len(values) else None
                    row[header] = "" if value is None else str(value)
                rows.append(row)
            return rows
        finally:
            workbook.close()

    def _get_openpyxl_load_workbook(self):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("xlsx_parser_unavailable") from exc
        return load_workbook

    def _assert_xlsx_parser_available(self) -> None:
        self._get_openpyxl_load_workbook()

    def _parse_pdf(self, file_bytes: bytes) -> list[dict[str, str]]:
        reader = PdfReader(io.BytesIO(file_bytes))
        lines: list[str] = []
        row_count = 0
        for page in reader.pages:
            text_value = page.extract_text() or ""
            page_lines = [line.strip() for line in text_value.splitlines() if line.strip()]
            lines.extend(page_lines)
            row_count += len(page_lines)
            if row_count > MAX_IMPORT_ROWS:
                raise ParserLimitError("max_rows_exceeded")
        return self._parse_text_lines(lines)

    def _parse_delimited_text(self, text_value: str) -> list[dict[str, str]]:
        sample = text_value[:1024]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.get_dialect("excel")

        reader = csv.DictReader(io.StringIO(text_value), dialect=dialect)
        rows: list[dict[str, str]] = []
        cell_count = 0
        for row_count, row in enumerate(reader, start=1):
            if row_count > MAX_IMPORT_ROWS:
                raise ParserLimitError("max_rows_exceeded")
            cell_count += len(row)
            if cell_count > MAX_IMPORT_CELLS:
                raise ParserLimitError("max_cells_exceeded")
            rows.append({str(k): "" if v is None else str(v) for k, v in row.items()})
        return rows

    def _parse_text_lines(self, lines: list[str]) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        current_location_name = ""
        current_city = ""
        current_state = ""
        current_category = ""

        for line in lines:
            if ":" in line:
                prefix, value = line.split(":", 1)
                key = prefix.strip().casefold()
                content = value.strip()
                if key in {"location", "site", "plant", "location name"}:
                    current_location_name = content
                    continue
                if key == "city":
                    current_city = content
                    continue
                if key == "state":
                    current_state = content
                    continue
                if key == "category":
                    current_category = content
                    continue
                if key in {"project", "waste stream", "stream"}:
                    rows.append(
                        {
                            "location_name": current_location_name,
                            "city": current_city,
                            "state": current_state,
                            "project_name": content,
                            "category": current_category,
                        }
                    )
                    if len(rows) > MAX_IMPORT_ROWS:
                        raise ParserLimitError("max_rows_exceeded")
                    continue

            comma_parts = [part.strip() for part in line.split(",")]
            if len(comma_parts) >= 2:
                row: dict[str, str] = {
                    "project_name": comma_parts[0],
                    "category": comma_parts[1],
                }
                if len(comma_parts) >= 5:
                    row["location_name"] = comma_parts[2]
                    row["city"] = comma_parts[3]
                    row["state"] = comma_parts[4]
                rows.append(row)
                if len(rows) > MAX_IMPORT_ROWS:
                    raise ParserLimitError("max_rows_exceeded")

        return rows

    def _extract_location_data(self, row: dict[str, str | None]) -> dict[str, str] | None:
        name = self._pick_value(
            row, ["location_name", "location", "site", "plant", "location name"]
        )
        city = self._pick_value(row, ["city", "location_city"])
        state_value = self._pick_value(row, ["state", "province", "location_state"])
        address = self._pick_value(row, ["address", "location_address"])
        if not any([name, city, state_value, address]):
            return None
        return {
            "name": name,
            "city": city,
            "state": state_value,
            "address": address,
        }

    def _extract_project_data(self, row: dict[str, str | None]) -> dict[str, str] | None:
        name = self._pick_value(
            row, ["project_name", "waste_stream", "waste stream", "project", "name"]
        )
        category = self._pick_value(row, ["category", "waste_category", "waste category"])
        project_type = self._pick_value(row, ["project_type", "project type"]) or "Assessment"
        description = self._pick_value(row, ["description", "details"])
        sector = self._pick_value(row, ["sector"])
        subsector = self._pick_value(row, ["subsector", "sub_sector"])
        estimated_volume = self._pick_value(row, ["estimated_volume", "volume", "estimated volume"])

        if not any([name, category, description, estimated_volume]):
            return None

        return {
            "name": name,
            "category": category,
            "project_type": project_type,
            "description": description,
            "sector": sector,
            "subsector": subsector,
            "estimated_volume": estimated_volume,
        }

    def _pick_value(self, row: dict[str, str | None], keys: list[str]) -> str:
        for key in keys:
            value = row.get(key)
            if value and value.strip():
                return value.strip()
        return ""

    async def _build_import_items(
        self,
        db: AsyncSession,
        run: ImportRun,
        parsed_rows: list[ParsedRow],
    ) -> list[ImportItem]:
        if not parsed_rows:
            return []

        if run.entrypoint_type == "company":
            return await self._build_company_entrypoint_items(db, run, parsed_rows)
        return await self._build_location_entrypoint_items(db, run, parsed_rows)

    async def _build_company_entrypoint_items(
        self,
        db: AsyncSession,
        run: ImportRun,
        parsed_rows: list[ParsedRow],
    ) -> list[ImportItem]:
        company = await db.get(Company, run.entrypoint_id)
        if not company or company.organization_id != run.organization_id:
            raise ValueError("entrypoint_company_not_found")

        existing_locations_result = await db.execute(
            select(Location).where(
                Location.organization_id == run.organization_id,
                Location.company_id == company.id,
            )
        )
        existing_locations = existing_locations_result.scalars().all()
        existing_location_ids_by_key: dict[str, list[UUID]] = {}
        for existing in existing_locations:
            location_key = self._location_key(
                {
                    "name": existing.name,
                    "city": existing.city,
                    "state": existing.state,
                }
            )
            existing_location_ids_by_key.setdefault(location_key, []).append(existing.id)

        location_defs: dict[str, dict[str, Any]] = {}
        project_defs: list[dict[str, Any]] = []

        for row in parsed_rows:
            location_key: str | None = None
            if row.location_data is not None:
                normalized_location = self._normalize_location_payload(row.location_data)
                location_key = self._location_key(normalized_location)
                location_confidence = self._confidence_from_raw(row.raw, "location_confidence", 80)
                if location_key not in location_defs:
                    location_defs[location_key] = {
                        "normalized_data": normalized_location,
                        "raw": _sanitize_payload(row.raw),
                        "confidence": location_confidence,
                    }
                else:
                    existing_confidence = int(location_defs[location_key].get("confidence", 80))
                    location_defs[location_key]["confidence"] = max(
                        existing_confidence,
                        location_confidence,
                    )

            if row.project_data is not None:
                project_defs.append(
                    {
                        "normalized_data": self._normalize_project_payload(row.project_data),
                        "raw": _sanitize_payload(row.raw),
                        "location_key": location_key,
                        "confidence": self._confidence_from_raw(row.raw, "stream_confidence", 75),
                    }
                )

        items: list[ImportItem] = []
        location_items_by_key: dict[str, ImportItem] = {}
        candidate_location_ids_by_key: dict[str, list[UUID]] = {}

        for key, payload in location_defs.items():
            normalized = payload["normalized_data"]
            duplicate_candidates = self._find_location_duplicates(existing_locations, normalized)
            candidate_location_ids_by_key[key] = self._candidate_location_ids_for_key(
                location_key=key,
                duplicate_candidates=duplicate_candidates,
                existing_location_ids_by_key=existing_location_ids_by_key,
            )
            confidence = int(payload.get("confidence", 80))
            needs_review = (
                self._needs_review("location", normalized)
                or bool(duplicate_candidates)
                or self._needs_review_by_confidence(confidence)
            )
            item = ImportItem(
                organization_id=run.organization_id,
                run_id=run.id,
                item_type="location",
                status="pending_review",
                needs_review=needs_review,
                confidence=confidence,
                extracted_data=payload["raw"],
                normalized_data=normalized,
                duplicate_candidates=duplicate_candidates or None,
                confirm_create_new=False,
            )
            items.append(item)
            location_items_by_key[key] = item

        if items:
            db.add_all(items)
            await db.flush()

        all_candidate_location_ids: list[UUID] = []
        seen_location_ids: set[UUID] = set()
        for location_ids in candidate_location_ids_by_key.values():
            for location_id in location_ids:
                if location_id in seen_location_ids:
                    continue
                seen_location_ids.add(location_id)
                all_candidate_location_ids.append(location_id)
        project_index_by_location_and_name = await self._prefetch_project_index_for_locations(
            db=db,
            organization_id=run.organization_id,
            location_ids=all_candidate_location_ids,
        )

        for payload in project_defs:
            normalized = payload["normalized_data"]
            location_key = payload["location_key"]
            parent_item = location_items_by_key.get(location_key) if location_key else None

            if parent_item is None:
                confidence = int(payload.get("confidence", 50))
                item = ImportItem(
                    organization_id=run.organization_id,
                    run_id=run.id,
                    item_type="project",
                    status="invalid",
                    needs_review=False,
                    confidence=confidence,
                    extracted_data=payload["raw"],
                    normalized_data=normalized,
                    review_notes="Project row missing location context",
                    confirm_create_new=False,
                )
                items.append(item)
                continue

            location_ids = (
                candidate_location_ids_by_key.get(location_key, []) if location_key else []
            )
            duplicate_candidates = self._match_project_duplicates_from_index(
                project_index_by_location_and_name=project_index_by_location_and_name,
                location_ids=location_ids,
                project_data=normalized,
            )
            confidence = int(payload.get("confidence", 75))
            needs_review = (
                self._needs_review("project", normalized)
                or bool(duplicate_candidates)
                or self._needs_review_by_confidence(confidence)
            )
            item = ImportItem(
                organization_id=run.organization_id,
                run_id=run.id,
                item_type="project",
                status="pending_review",
                needs_review=needs_review,
                confidence=confidence,
                extracted_data=payload["raw"],
                normalized_data=normalized,
                duplicate_candidates=duplicate_candidates or None,
                parent_item_id=parent_item.id if parent_item else None,
                confirm_create_new=False,
            )
            items.append(item)

        return items

    async def _build_location_entrypoint_items(
        self,
        db: AsyncSession,
        run: ImportRun,
        parsed_rows: list[ParsedRow],
    ) -> list[ImportItem]:
        location = await db.get(Location, run.entrypoint_id)
        if not location or location.organization_id != run.organization_id:
            raise ValueError("entrypoint_location_not_found")

        project_index_by_location_and_name = await self._prefetch_project_index_for_locations(
            db=db,
            organization_id=run.organization_id,
            location_ids=[location.id],
        )

        items: list[ImportItem] = []
        invalid_locations_by_key: dict[str, ImportItem] = {}

        for row in parsed_rows:
            if row.location_data:
                normalized_location = self._normalize_location_payload(row.location_data)
                key = self._location_key(normalized_location)
                if key and key not in invalid_locations_by_key:
                    confidence = self._confidence_from_raw(row.raw, "location_confidence", 50)
                    invalid_item = ImportItem(
                        organization_id=run.organization_id,
                        run_id=run.id,
                        item_type="location",
                        status="invalid",
                        needs_review=False,
                        confidence=confidence,
                        extracted_data=_sanitize_payload(row.raw),
                        normalized_data=normalized_location,
                        review_notes="Location items invalid for location entrypoint",
                    )
                    invalid_locations_by_key[key] = invalid_item
                    items.append(invalid_item)

            if not row.project_data:
                continue

            normalized_project = self._normalize_project_payload(row.project_data)
            is_external_location = False
            if row.location_data:
                normalized_location = self._normalize_location_payload(row.location_data)
                is_external_location = self._is_external_location(location, normalized_location)

            duplicate_candidates = self._match_project_duplicates_from_index(
                project_index_by_location_and_name=project_index_by_location_and_name,
                location_ids=[location.id],
                project_data=normalized_project,
            )
            status_value = "invalid" if is_external_location else "pending_review"
            review_notes = (
                "Project row references external location" if is_external_location else None
            )
            confidence = self._confidence_from_raw(row.raw, "stream_confidence", 75)
            needs_review = (
                False
                if status_value == "invalid"
                else (
                    self._needs_review("project", normalized_project)
                    or bool(duplicate_candidates)
                    or self._needs_review_by_confidence(confidence)
                )
            )

            item = ImportItem(
                organization_id=run.organization_id,
                run_id=run.id,
                item_type="project",
                status=status_value,
                needs_review=needs_review,
                confidence=confidence,
                extracted_data=_sanitize_payload(row.raw),
                normalized_data=normalized_project,
                duplicate_candidates=duplicate_candidates or None,
                review_notes=review_notes,
                confirm_create_new=False,
            )
            items.append(item)

        if invalid_locations_by_key:
            db.add_all(list(invalid_locations_by_key.values()))
            await db.flush()

        return items

    def _normalize_location_payload(self, payload: dict[str, str]) -> dict[str, object]:
        normalized = {
            "name": _sanitize_text(payload.get("name") or "") or "",
            "city": _sanitize_text(payload.get("city") or "") or "",
            "state": _sanitize_text(payload.get("state") or "") or "",
            "address": _sanitize_text(payload.get("address") or ""),
        }
        return _sanitize_payload(normalized)

    def _normalize_project_payload(self, payload: dict[str, str]) -> dict[str, object]:
        normalized = {
            "name": _sanitize_text(payload.get("name") or "") or "",
            "category": _sanitize_text(payload.get("category") or ""),
            "project_type": _sanitize_text(payload.get("project_type") or "Assessment")
            or "Assessment",
            "description": _sanitize_text(payload.get("description") or ""),
            "sector": _sanitize_text(payload.get("sector") or ""),
            "subsector": _sanitize_text(payload.get("subsector") or ""),
            "estimated_volume": _sanitize_text(payload.get("estimated_volume") or ""),
        }
        return _sanitize_payload(normalized)

    def _location_key(self, normalized_location: dict[str, object]) -> str:
        name = _normalize_token(str(normalized_location.get("name") or ""))
        city = _normalize_token(str(normalized_location.get("city") or ""))
        state_value = _normalize_token(str(normalized_location.get("state") or ""))
        return f"{name}|{city}|{state_value}"

    def _find_location_duplicates(
        self,
        existing_locations: Sequence[Location],
        normalized_location: dict[str, object],
    ) -> list[dict[str, object]]:
        name = _normalize_token(str(normalized_location.get("name") or ""))
        city = _normalize_token(str(normalized_location.get("city") or ""))
        state_value = _normalize_token(str(normalized_location.get("state") or ""))
        if not name:
            return []

        candidates: list[dict[str, object]] = []
        for existing in existing_locations:
            existing_name = _normalize_token(existing.name)
            if existing_name != name:
                continue

            reasons = ["name_match"]
            if city and state_value:
                if (
                    _normalize_token(existing.city) == city
                    and _normalize_token(existing.state) == state_value
                ):
                    reasons.extend(["city_match", "state_match"])
                else:
                    continue
            candidates.append(
                {
                    "id": str(existing.id),
                    "name": existing.name,
                    "reason_codes": reasons,
                }
            )
        return candidates

    def _candidate_location_ids_for_key(
        self,
        *,
        location_key: str,
        duplicate_candidates: list[dict[str, object]],
        existing_location_ids_by_key: dict[str, list[UUID]],
    ) -> list[UUID]:
        candidate_ids: list[UUID] = []
        for candidate in duplicate_candidates:
            raw_id = candidate.get("id")
            if not isinstance(raw_id, str):
                continue
            try:
                candidate_ids.append(UUID(raw_id))
            except ValueError:
                continue

        if not candidate_ids:
            candidate_ids.extend(existing_location_ids_by_key.get(location_key, []))

        deduped: list[UUID] = []
        seen: set[UUID] = set()
        for candidate_id in candidate_ids:
            if candidate_id in seen:
                continue
            seen.add(candidate_id)
            deduped.append(candidate_id)
        return deduped

    async def _prefetch_project_index_for_locations(
        self,
        *,
        db: AsyncSession,
        organization_id: UUID,
        location_ids: Sequence[UUID],
    ) -> dict[tuple[UUID, str], list[Project]]:
        unique_location_ids = list(dict.fromkeys(location_ids))
        if not unique_location_ids:
            return {}

        result = await db.execute(
            select(Project).where(
                Project.organization_id == organization_id,
                Project.location_id.in_(unique_location_ids),
            )
        )
        projects = result.scalars().all()
        indexed: dict[tuple[UUID, str], list[Project]] = {}
        for project in projects:
            if project.location_id is None:
                continue
            normalized_name = _normalize_token(project.name)
            if not normalized_name:
                continue
            index_key = (project.location_id, normalized_name)
            if index_key not in indexed:
                indexed[index_key] = []
            indexed[index_key].append(project)
        return indexed

    def _match_project_duplicates_from_index(
        self,
        *,
        project_index_by_location_and_name: dict[tuple[UUID, str], list[Project]],
        location_ids: Sequence[UUID],
        project_data: dict[str, object],
    ) -> list[dict[str, object]]:
        project_name = _normalize_token(str(project_data.get("name") or ""))
        if not project_name:
            return []

        candidates: list[dict[str, object]] = []
        seen_project_ids: set[UUID] = set()
        for location_id in location_ids:
            matching_projects = project_index_by_location_and_name.get(
                (location_id, project_name), []
            )
            for project in matching_projects:
                if project.id in seen_project_ids:
                    continue
                seen_project_ids.add(project.id)
                candidates.append(
                    {
                        "id": str(project.id),
                        "name": project.name,
                        "reason_codes": ["name_match", "location_match"],
                    }
                )
        return candidates

    def _is_external_location(
        self, entrypoint_location: Location, candidate: dict[str, object]
    ) -> bool:
        candidate_name = _normalize_token(str(candidate.get("name") or ""))
        candidate_city = _normalize_token(str(candidate.get("city") or ""))
        candidate_state = _normalize_token(str(candidate.get("state") or ""))
        current_name = _normalize_token(entrypoint_location.name)
        current_city = _normalize_token(entrypoint_location.city)
        current_state = _normalize_token(entrypoint_location.state)

        if candidate_name and candidate_name != current_name:
            return True
        if candidate_city and candidate_city != current_city:
            return True
        return bool(candidate_state and candidate_state != current_state)
